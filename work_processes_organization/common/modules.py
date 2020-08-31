import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import datetime
from os.path import isfile, join, splitext, exists


class ProjectScheduler:
    """
    Этот класс содержит методы которые собирают данные с таблиц и добавляет их в базу данных.
    Если не задать аргументы метод возвращает список файлов .xlsx с папки SCHEDULES
    Если передать аргумент filename (должен соответствовать существующим файлам)
    то скопирует содержимое файла в базу данных.
    """

    root_folder = r"\\mcp-fsvs2\Schedules"
    # def __init__(self, job, job_name, release, pm, pt_dwg=None, zha=None, flashing=None, coping=None,
    #              splice_pate=None, blade_screen=None, perf=None, plate_panels=None, frames=None, strapping=None,
    #              misc=None, est_mh=None, rel_date=None, shipped_date=None, date_dif=None, status=None, shipped_to=None,
    #              remarks=None, loose_item=False, out_paint=False):
    #     self.job = job
    #     self.job_name = job_name
    #     self.release = release
    #     self.pm = pm
    #     self.loose_item = loose_item
    #     self.out_paint = out_paint
    #     self.pt_dwg = pt_dwg
    #     self.zha = zha
    #     self.flashing = flashing
    #     self.coping = coping
    #     self.splice_pate = splice_pate
    #     self.blade_screen = blade_screen
    #     self.perf = perf
    #     self.plate_panels = plate_panels
    #     self.frames = frames
    #     self.strapping = strapping
    #     self.misc = misc
    #     self.est_mh = est_mh
    #     self.rel_date = rel_date
    #     self.shipped_date = shipped_date
    #     self.date_dif = date_dif
    #     self.status = status
    #     self.shipped_to = shipped_to
    #     self.remarks = remarks

    @staticmethod
    def get_book(src_dir=root_folder, filename=None):
        """
        get book if exist otherwise return actual file list.
        """
        if filename:
            try:  # пробуем сопоставить строку и реальным именем файла, если файл найден то читаем все в переменную wb
                filepath = join(ProjectScheduler.root_folder, filename)
                print(filepath)
                wb = openpyxl.load_workbook(filepath)
                return wb  # возвращаем отдельно 2 листа книги
            except FileNotFoundError:
                print(f'Couldn\'t open "{filename}" file not found')
        else:  # если не задано имя файла то возвращает список всех Эксель файлов в директории.
            files = os.listdir(src_dir)
            return [x for x in files if isfile(join(src_dir, x))
                    and x.endswith('.xlsx')
                    or x.endswith('.csv')
                    and not x.startswith('~')]

    @staticmethod
    def export_to_file():
        """
        TODO: export_to_file()
        """
        pass

    @staticmethod
    def get_pro_list(sheet_job):  # as {Job_number:[title, saleman, PM]}
        """ func -> dict() This dict important to find Project Title in the Excel file """
        # sheet_job = sheet_job
        job_dict = dict()
        for i in range(2, sheet_job.max_row):
            lst = []
            for j in range(1, 5):
                lst.append(sheet_job.cell(row=i, column=j).value)
                # print(lst)
            job_dict.update({lst.pop(0): lst})
        # for x in job_dict:
        #     print(f'{x}: {job_dict[x]}')
        return job_dict

    @staticmethod
    def define_line_type(row_num):
        """
        DESCRIPTION: this func defines line type, it is important for further handling.
        :return Str
        """
        if sheet.cell(row=row_num, column=1).value in ['Job', 'Job #']:  # Headers
            line_type = 'HEADERS'
            return line_type
        elif isinstance(sheet.cell(row=row_num, column=5).value, datetime.datetime):  # Datetime
            line_type = 'DATETIME'
            return line_type
        elif isinstance(sheet.cell(row=row_num, column=1).value, int):
            line_type = "DATA"
            return line_type
        elif isinstance(sheet.cell(row=row_num, column=7).value, str) \
                and sheet.cell(row=row_num, column=7).value.startswith('='):
            return 'SUMMARY LINE'
        else:
            return 'EMPTY LINE'

    @staticmethod
    def get_task_list():
        """
        TODO: get_task_list
        """
        pass

    @staticmethod
    def import_data_to_db(row_num):
        """
        DESCRIPTION: import data to DB
        :var row_num of active sheet
        :return None
        TODO: Import Excel management data to DB.Tasks
        """
        pass


book = ProjectScheduler.get_book(filename='Metal Shop Schedule - 2020.xlsx')
sheet = book.get_sheet_by_name('Sheet1')
job_sheet = book.get_sheet_by_name('Job Specific ')
proj_list = ProjectScheduler.get_pro_list(job_sheet)

""" 
с порядком тут какие-то проблемы, не знаю почему мне пришлось поменять местами переменные, так-что 
второй лист копируется в первую переменную, а 
первый лист во вторую переменную
через время все прошло само по себе О_о
Решением было явно прописывать названия Листов. (может аукнуться в будущем)
"""
# print(ProjectScheduler.get_pro_list(job_sheet))

""" 
*****************
*** TEST AREA *** 
*****************
"""

print(f'rows = {sheet.max_row}, columns = {sheet.max_column}')  # statistic data
for row in range(1, sheet.max_row):  # sheet.max_row
    # print(f'\nLine:{row} -= {ExcelHandler.define_line_type(row)} =-\n')
    for cell in range(1, sheet.max_column):
        cell_value = sheet.cell(row=row, column=cell).value
        if cell_value:
            if isinstance(cell_value, (str, int)):  # and not cell_value.startswith('=')
                if cell == 2 and ProjectScheduler.define_line_type(row) == "DATA":
                    repl = sheet.cell(row=row, column=1).value
                    try:
                        print(proj_list[repl][0])
                    except KeyError:
                        continue
                elif ProjectScheduler.define_line_type(row) in ["HEADERS", "SUMMARY LINE", "EMPTY LINE", "DATETIME"]:
                    continue
                else:
                    print(sheet.cell(row=row, column=cell).value)

