import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import datetime

wb = openpyxl.load_workbook(r'automate_online-materials/example.xlsx')
# 1
# print(type(wb))
# print(wb.sheetnames)
# print(len(wb.sheetnames))
# sheet = wb[wb.sheetnames[-1]]
# # sheet = wb['Sheet3']
# print(type(sheet))
# print(sheet.title)
# anotherSheet = wb.active
# print(anotherSheet)
#
# 2
sheet = wb.active
print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
print('Cell %s is %s' % (c.coordinate, c.value))
print(sheet['C1'].value)
#
# 3
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

print(sheet.max_row, sheet.max_column)
print(get_column_letter(18278))
print(get_column_letter(sheet.max_column))
print(column_index_from_string('ZZZ'))
